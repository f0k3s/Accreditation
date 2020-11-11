import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

from docxtpl import DocxTemplate
#Функция поиска и вывода строки из экселя
def search (x1):
    if (x1 < a) and (x1 > 0):
        x1 = x1-1
        return (dict['a1'][x1],dict['a2'][x1],dict['a3'][x1],dict['a4'][x1],dict['a5'][x1])
    else:
        return ('Такого нет!')
#Функция записывающая столбец в элемент
def sbor (x, y, z):
    for cellObj in sheet[ y : z]:
      for cell in cellObj:
          x.append(cell.value)

#Начинаем создавать файл
wb = Workbook()
ws = wb.active

ws['A1'] = '№'
ws['B1'] = 'ФИО'
ws['C1'] = 'Предмет'
ws['D1'] = 'Уч.степень'
ws['E1'] = 'Кабинет'
ws.append([1, 'Пётр Перов', 'информатика', 'Доцент', 'А262'])
ws.append([2, 'Владимир Владимирович', 'бокс', 'Старший преподаватель', 'Б104'])
ws.append([3, 'ФИО', 'физика', 'Доцент', 'В303'])
ws.append([4, 'Человек человеков', 'математика', 'Профессор', 'Г415'])
ws.append([5, 'Кто живой', 'жизнь', 'Доцент', 'Д101'])

#Сохраняем файл
wb.save("sample.xlsx")
wb.close()

#Загружаем файл
wb = load_workbook('sample.xlsx')
sheet = wb.get_sheet_by_name('Sheet')
a = int(input('Сколько строк? '))
a = a + 1
b = 'E' + str(a)
#Вывести на экран
x=int(input('Вывести значения на экран?(Введите 1):'))
if x == 1:
    for cellObj in sheet['A1': b]:
        for cell in cellObj:
            print(cell.value)
        print('--- Конец строки ---')

#Сбор данных по отдельным массивам словарь
dict = {'a1':[], 'a2':[], 'a3':[], 'a4':[], 'a5':[]}
ba = 'A' + str(a)
sbor(dict['a1'], 'A2', ba)
ba = 'B' + str(a)
sbor(dict['a2'], 'B2', ba)
ba = 'C' + str(a)
sbor(dict['a3'], 'C2', ba)
ba = 'D' + str(a)
sbor(dict['a4'], 'D2', ba)
ba = 'E' + str(a)
sbor(dict['a5'], 'E2', ba)

#Перенос текста в ворд в определённые места
doc = DocxTemplate("шаблон.docx")
for i in range(a):
    context = { 'FIO' : dict['a2'][1], 'Predm' : dict['a3'][1], 'Stepen' : dict['a4'][1], 'Kab' : dict

['a5'][1]}
    doc.render(context)
    doc.save("шаблон-final.docx")

#Вывод определённого столбца на экран 

z=1
while z != 0:
    print('Выберите какой столбик вывести на экран: ')
    print('1 – №')
    print('2 – ФИО')
    print('3 – Предмет')
    print('4 – Уч.степень')
    print('5 – Кабинет')
    x = int(input('Ваш выбор – '))
    if  1<=x<=5:
               if x==1:
                    print(dict['a1'])
                    x1 = int(input('Вывести конкретного?(цифра) '))
                    print (search(x1))
               if x==2:
                    print(dict['a2'])
                    x1 = int(input('Вывести конкретного?(цифра) '))
                    print (search(x1))
               if x==3:
                    print(dict['a3'])
                    x1 = int(input('Вывести конкретного?(цифра) '))
                    print (search(x1))
               if x==4:
                    print(dict['a4'])
                    x1 = int(input('Вывести конкретного?(цифра) '))
                    print (search(x1))
               if x==5:
                    print(dict['a5'])
                    x1 = int(input('Вывести конкретного?(цифра) '))
                    print (search(x1))
    else:
       print('Выбран не верный вариант!')
    print('Для продолжения нажмите 1, иначе будет выход из программы')
    b=int(input())
    if b!=1:
        z-=1


# Создание таблицы в ворде
#import docx
#doc = docx.Document('шаблон.docx')
#table = doc.add_table(rows = a, cols = 4)
#table.style = 'Table Grid'
#aa = a-1
#for row in range(aa):
#    for col in range(1):
#        # получаем ячейку таблицы
#        cell = table.cell(row, col)
#        # записываем в ячейку данные
#        cell.text =  ???     #Надо продумать заполнение таблицы
#doc.save('table.docx')
